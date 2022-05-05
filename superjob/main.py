import openpyxl
from openpyxl.worksheet import worksheet

from superjob import url_scraping, excel_writing
from superjob.excel_writing import URL_XLSX_FILENAME


def add_url():
    book = openpyxl.load_workbook(filename=URL_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    num = sheet.max_row
    urls = []
    for i in range(1, num + 1):
        link_cell = sheet.cell(row=i, column=1)
        urls.append(link_cell.value)
    return urls

if __name__ == '__main__':
    # urls = url_scraping.scrap_urls()
    # excel_writing.urls_to_xlsx(urls)
    urls = add_url()
    for url in urls:
        excel_writing.dump_to_xlsx(url)
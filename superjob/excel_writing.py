import openpyxl
import requests
from lxml import etree
from openpyxl.worksheet import worksheet
from bs4 import BeautifulSoup as BS
from superjob import data_scraping

OUT_XLSX_FILENAME = 'superjob_parsing.xlsx'
URL_XLSX_FILENAME = 'urls.xlsx'

def urls_to_xlsx(urls):
    book = openpyxl.load_workbook(filename=URL_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    num = sheet.max_row
    for url in urls:
        url = 'https://www.superjob.ru' + url
        sheet.cell(row=num, column=1, value=url)
        num += 1
    book.save(URL_XLSX_FILENAME)


def dump_to_xlsx(url):
    book = openpyxl.load_workbook(filename=OUT_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    html = requests.get(url)
    soup = BS(html.content, 'html.parser')
    num = sheet.max_row + 1
    sheet.cell(row=num, column=1, value=data_scraping.scrap_id(soup))
    sheet.cell(row=num, column=2, value=data_scraping.scrap_updated(soup))
    sheet.cell(row=num, column=3, value=data_scraping.scrap_category(soup))
    sheet.cell(row=num, column=4, value=data_scraping.scrap_online(soup))
    sheet.cell(row=num, column=5, value=data_scraping.scrap_merried(soup))
    sheet.cell(row=num, column=6, value=data_scraping.scrap_child(soup))
    sheet.cell(row=num, column=7, value=data_scraping.scrap_sex(soup))
    sheet.cell(row=num, column=8, value=data_scraping.scrap_position(soup))
    sheet.cell(row=num, column=9, value=data_scraping.scrap_employment(soup))
    sheet.cell(row=num, column=10, value=data_scraping.scrap_birth(soup))
    sheet.cell(row=num, column=11, value=data_scraping.scrap_age(soup))
    sheet.cell(row=num, column=14, value=data_scraping.scrap_trip(soup))
    sheet.cell(row=num, column=15, value=data_scraping.scrap_live(soup))
    sheet.cell(row=num, column=19, value=data_scraping.scrap_vaccine(soup))
    sheet.cell(row=num, column=20, value=data_scraping.scrap_salary(soup))
    sheet.cell(row=num, column=21, value=data_scraping.scrap_curr(soup))
    sheet.cell(row=num, column=22, value=data_scraping.scrap_citiz(soup))
    sheet.cell(row=num, column=23, value=data_scraping.scrap_exp(soup))
    sheet.cell(row=num, column=24, value=data_scraping.scrap_wempl(soup))
    sheet.cell(row=num, column=25, value=data_scraping.scrap_org(soup))
    sheet.cell(row=num, column=26, value=data_scraping.scrap_sitec(soup))
    sheet.cell(row=num, column=29, value=data_scraping.scrap_startw(soup))
    sheet.cell(row=num, column=30, value=data_scraping.scrap_endw(soup))
    sheet.cell(row=num, column=31, value=data_scraping.scrap_long(soup))
    sheet.cell(row=num, column=32, value=data_scraping.scrap_duties(soup))
    sheet.cell(row=num, column=34, value=data_scraping.scrap_lvl(soup))
    sheet.cell(row=num, column=35, value=data_scraping.scrap_form(soup))
    sheet.cell(row=num, column=36, value=data_scraping.scrap_year(soup))
    sheet.cell(row=num, column=37, value=data_scraping.scrap_name(soup))
    sheet.cell(row=num, column=38, value=data_scraping.scrap_fac(soup))
    sheet.cell(row=num, column=39, value=data_scraping.scrap_spec(soup))
    sheet.cell(row=num, column=44, value=data_scraping.scrap_prof(soup))
    sheet.cell(row=num, column=45, value=data_scraping.scrap_add(soup))
    sheet.cell(row=num, column=46, value=data_scraping.scrap_lan(soup))
    sheet.cell(row=num, column=47, value=data_scraping.scrap_lvlknw(soup))
    sheet.cell(row=num, column=48, value=data_scraping.scrap_drive(soup))
    sheet.cell(row=num, column=49, value=url)
    book.save(OUT_XLSX_FILENAME)



























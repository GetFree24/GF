import kwork_parsing
import openpyxl
from openpyxl.worksheet import worksheet
from openpyxl.styles import Font


OUT_XLSX_FILENAME = 'kwork_parsing.xlsx'

def dump_to_xlsx():
    book = openpyxl.load_workbook(filename=OUT_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    sheet.cell(row=1, column=1, value='Категории')
    sheet.cell(row=1, column=2, value='Название')
    sheet.cell(row=1, column=3, value='Желаемая цена')
    sheet.cell(row=1, column=4, value='Допустимая цена')
    sheet.cell(row=1, column=5, value='Время до окнчания приема заявок')
    sheet.cell(row=1, column=6, value='Количество предложений')
    sheet.cell(row=1, column=7, value='Ник заказчика')
    sheet.cell(row=1, column=8, value='Нанято')
    sheet.cell(row=1, column=9, value='Размещено проектов на бирже')
    sheet.cell(row=1, column=10, value='Описание')
    book.save(OUT_XLSX_FILENAME)



import openpyxl
from openpyxl.worksheet import worksheet
import habr_scraper

OUT_XLSX_FILENAME = 'habr_parsing.xlsx'
URL_XLSX_FILENAME = 'urls.xlsx'


def add_url():
    book = openpyxl.load_workbook(filename=URL_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    num = sheet.max_row
    urls = []
    for i in range(1, num + 1):
        link_cell = sheet.cell(row=i, column=1)
        urls.append(link_cell.value)
    return urls

def urls_to_xlsx(urls):
    book = openpyxl.load_workbook(filename=URL_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    num = 1
    for url in urls:
        sheet.cell(row=num, column=1, value=url)
        num += 1
    book.save(URL_XLSX_FILENAME)

def dump_to_xlsx():
    book = openpyxl.load_workbook(filename=OUT_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    sheet.cell(row=1, column=1, value='Категории')
    sheet.cell(row=1, column=2, value='Название')
    sheet.cell(row=1, column=3, value='Цена')
    sheet.cell(row=1, column=4, value='Валюта')
    sheet.cell(row=1, column=5, value='Дата публикации')
    sheet.cell(row=1, column=6, value='Количество откликов')
    sheet.cell(row=1, column=7, value='Количество предложений')
    sheet.cell(row=1, column=8, value='Описание')
    book.save(filename=OUT_XLSX_FILENAME)

def data_to_xlsx(soup, num):
    book = openpyxl.load_workbook(filename=OUT_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    sheet.cell(row=num, column=1, value=habr_scraper.scrap_category(soup))
    sheet.cell(row=num, column=2, value=habr_scraper.scrap_title(soup))
    sheet.cell(row=num, column=3, value=habr_scraper.scrap_price(soup))
    sheet.cell(row=num, column=4, value=habr_scraper.scrap_currency(soup))
    sheet.cell(row=num, column=5, value=habr_scraper.scrap_meta(soup)[0])
    sheet.cell(row=num, column=6, value=habr_scraper.scrap_meta(soup)[1])
    sheet.cell(row=num, column=7, value=habr_scraper.scrap_meta(soup)[2])
    sheet.cell(row=num, column=8, value=habr_scraper.scrap_description(soup))
    book.save(filename=OUT_XLSX_FILENAME)

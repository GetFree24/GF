import re

import openpyxl
from bs4 import BeautifulSoup as BS
import requests
from openpyxl.worksheet import worksheet


def scrap_urls(soup):
    urls = []
    for url_code in soup.find_all('div', class_='task__title'):
        url = url_code.find('a')
        urls.append('https://freelance.habr.com' + url['href'])
    return urls

def scrap_category(soup):
    category = soup.find('a', class_='tags__item_link')
    if category:
        return category.text

def scrap_title(soup):
    title = soup.find('h2', class_='task__title')
    return title.text

def scrap_price(soup):
    price_code = soup.find('span', class_='count')
    if price_code:
        price = re.split('[а-яёА-ЯЁ]+', price_code.text)

        return price[0]

def scrap_currency(soup):
    currency_code = soup.find('span', class_='count')
    if currency_code:
        currency = re.search('[а-яёА-ЯЁ]+', currency_code.text)
        return currency.group(0)

def scrap_description(soup):
    description = soup.find('div', class_='task__description')
    return description.text

def scrap_meta(soup):
    meta = soup.find('div', class_='task__meta')
    meta = meta.text.replace('\n', '')
    meta = re.split('•', meta)
    return meta







# def scrap_pages(soup):
#     max_page = 2
#     try:
#         pages = []
#         pagination = soup.find('div', class_='pagination')
#         for page_code in pagination.find_all('a'):
#             if page_code.text.isdigit():
#                 pages.append(page_code.text)
#         max_page = int(pages[len(pages) - 1])
#
#     except Exception:
#         pass
#     finally:
#         return max_page
#
#
# def data_scraping(soup, url):
#     book = openpyxl.load_workbook(filename=excel_writing.OUT_XLSX_FILENAME)
#     sheet: worksheet = book.worksheets[0]
#     num = sheet.max_row + 1
#     for data_code in soup.find_all('article', class_='task task_list'):
#         sheet.cell(row=num, column=1, value=category_finder(url))
#         titles_code = data_code.find('div', class_='task__title')
#         title_code = titles_code.find('a')
#         sheet.cell(row=num, column=2, value=title_code.text)
#         price_code = data_code.find('span', class_='count')
#         if price_code:
#             price = re.search('\d+\s\d*\s\d*', price_code.text)
#             sheet.cell(row=num, column=3, value=price.group(0))
#
#             sheet.cell(row=num, column=4, value=price[1])
#         # else:
#         #     price_code = data_code.find('span', class_='negotiated_price')
#         #     sheet.cell(row=num, column=4, value=price_code.text)
#         count_calling = data_code.find('i', class_='params__count')
#         sheet.cell(row=num, column=5, value=count_calling.text)
#         num += 1
#     book.save(excel_writing.OUT_XLSX_FILENAME)
#


def category_finder(url):
    category = ''
    if url.find('development_all_inclusive') != -1:
        category = 'Разработка'
    elif url.find('testing_sites') != -1:
        category = 'Тестирование'
    elif url.find('admin_servers') != -1:
        category = 'Администрирование'
    elif url.find('design_sites') != -1:
        category = 'Дизайн'
    elif url.find('content_copywriting') != -1:
        category = 'Контент'
    elif url.find('marketing_smm') != -1:
        category = 'Маркетинг'
    elif url.find('other_audit_analytics') != -1:
        category = 'Разное'
    return category




import openpyxl
from openpyxl.worksheet import worksheet

import connector
import excel_writing
import kwork_parsing
import get_url
from bs4 import BeautifulSoup as BS

URLS_PATH = 'urls.xlsx'

def add_url():
    book = openpyxl.load_workbook(filename=URLS_PATH)
    sheet: worksheet = book.worksheets[0]
    num = sheet.max_row
    urls = []
    for i in range(1, num + 1):
        link_cell = sheet.cell(row=i, column=1)
        urls.append(link_cell.value)
    return urls

if __name__ == '__main__':
    excel_writing.dump_to_xlsx()
    book = openpyxl.load_workbook(filename=excel_writing.OUT_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    html = connector.undetected_connection(f'https://kwork.ru/projects')
    soup = BS(html, "html.parser")
    num = sheet.max_row
    urls = add_url()
    # get_url.final_urls(soup, num + 1)
    for url in urls:
        html = connector.undetected_connection(url)
        soup = BS(html, "html.parser")
        num = kwork_parsing.general_parsing(soup, num + 1)
    # html = connector.undetected_connection('https://kwork.ru/projects')
    # soup = BS(html, "html.parser")
    # get_url.scrap_page_count(soup)


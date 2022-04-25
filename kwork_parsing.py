import re

import openpyxl
from openpyxl.worksheet import worksheet

link = 'https://kwork.ru/projects'
OUT_XLSX_FILENAME = 'kwork_parsing.xlsx'



def general_parsing(soup, num):
    book = openpyxl.load_workbook(filename=OUT_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    all_category = ''
    for general_element in soup.find_all('div', class_='card__content pb5'):
        for category_code in soup.find_all('a', 'chosen-single'):
            for category in category_code.find_all('span'):
                all_category += category.text + '/'
            sheet.cell(row=num, column=1, value=all_category)
        for name_code in general_element.find_all('div', class_='wants-card__header-title first-letter breakwords pr250'):
            if name_code is None:
                name_code = general_element.find('div', 'wants-card__header-title first-letter breakwords pr200')
            # name = name_code.find('a')
            sheet.cell(row=num, column=2, value=name_code.text)
        for fprice_code in general_element.find_all('div', class_='wants-card__header-price wants-card__price m-hidden'):
            fprice = fprice_code.text
            fprice = fprice.replace('Желаемый бюджет: ', '')
            fprice = fprice.replace('Цена ', '')
            sheet.cell(row=num, column=3, value=fprice)
        for sprice_code in general_element.find_all('div', class_='wants-card__description-higher-price'):
            sprice = sprice_code.text
            sprice = sprice.replace('Допустимый: ', '')
            sheet.cell(row=num, column=4, value=sprice)
        for time_sugges_code in general_element.find_all('div', class_='force-font force-font--s12'):
            times = []
            for time_code in time_sugges_code.find('span'):
                times.append(time_code.text)
            try:
                sheet.cell(row=num, column=5, value=(' ' if times[0] is None else times[0]))
            except Exception:
                pass
        for time_sugges_code in general_element.find_all('div', class_='force-font force-font--s12'):
            sugges = []
            for sugges_code in time_sugges_code.find_all('span'):
                suggess = sugges_code.text
                suggess = re.search('\d+', suggess)
                sugges.append(suggess.group(0))
            try:
                sheet.cell(row=num, column=6, value=(' ' if sugges[1] is None else sugges[1]))
            except Exception:
                pass
        for castomer_name in general_element.find_all('a', 'v-align-t'):
            sheet.cell(row=num, column=7, value=castomer_name.text)
        for rating_code in general_element.find_all('span', 'd-flex align-items-center'):
            rating = rating_code.text
            rating = re.search('\d+[%]', rating)
            sheet.cell(row=num, column=8, value=rating.group(0))
        for projects_code in general_element.find_all('div', class_='dib v-align-t'):
            project = projects_code.text
            project = re.search('\d+', project)
            sheet.cell(row=num, column=9, value=project.group(0))
        for details_code in general_element.find_all('div', class_='breakwords first-letter js-want-block-toggle js-want-block-toggle-full hidden'):
            sheet.cell(row=num, column=10, value=details_code.text)
        num += 1
    book.save(OUT_XLSX_FILENAME)
    return num


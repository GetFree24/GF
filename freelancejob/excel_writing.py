import openpyxl
from openpyxl.worksheet import worksheet
from bs4 import BeautifulSoup as BS
import requests

from freelancejob import freelancejob_scraper

OUT_XLSX_FILENAME = 'freelancejob_data.xlsx'
URL_XLSX_FILENAME = 'urls.xlsx'

def urls_to_xlsx():
    book = openpyxl.load_workbook(filename=URL_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    num = sheet.max_row
    page = int(freelancejob_scraper.get_pages())
    for x in range(1, page + 1):
        url = f'https://www.freelancejob.ru/projects/p{x}/'
        res = requests.get(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:47.0) Gecko/20100101 Firefox/47.0'})
        soup = BS(res.text, 'html.parser')
        values = freelancejob_scraper.urls_scraping(soup)
        for value in values:
            sheet.cell(row=num, column=1, value=value)
            num += 1
    book.save(filename=URL_XLSX_FILENAME)

def get_urls():
    book = openpyxl.load_workbook(filename=URL_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    num = sheet.max_row
    urls = []
    for i in range(1, num + 1):
        link_cell = sheet.cell(row=i, column=1)
        urls.append(link_cell.value)
    return urls

def fill_title():
    book = openpyxl.load_workbook(filename=OUT_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    sheet.cell(row=1, column=1, value='Категории')
    sheet.cell(row=1, column=2, value='Название')
    sheet.cell(row=1, column=3, value='Цена')
    sheet.cell(row=1, column=4, value='Валюта')
    sheet.cell(row=1, column=5, value='Город')
    sheet.cell(row=1, column=6, value='Вид предложения')
    sheet.cell(row=1, column=7, value='Описание')
    book.save(filename=OUT_XLSX_FILENAME)

def dump_to_xlsx(urls):
    book = openpyxl.load_workbook(filename=OUT_XLSX_FILENAME)
    sheet: worksheet = book.worksheets[0]
    num = 1
    for url in urls:
        try:
            res = requests.get(url, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:47.0) Gecko/20100101 Firefox/47.0'})
            soup = BS(res.text, 'html.parser')
            sheet.cell(row=num, column=1, value=freelancejob_scraper.get_category(soup))
            sheet.cell(row=num, column=2, value=freelancejob_scraper.scrap_title(soup))
            sheet.cell(row=num, column=3, value=freelancejob_scraper.get_price(soup))
            sheet.cell(row=num, column=4, value=freelancejob_scraper.get_currency(soup))
            sheet.cell(row=num, column=5, value=freelancejob_scraper.get_town(soup))
            sheet.cell(row=num, column=6, value=freelancejob_scraper.get_offer(soup))
            sheet.cell(row=num, column=7, value=freelancejob_scraper.get_details(soup))
            num += 1
        except Exception:
            pass
        finally:
            book.save(filename=OUT_XLSX_FILENAME)